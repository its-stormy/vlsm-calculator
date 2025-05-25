import ipaddress
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
import datetime

def calculate_vlsm(base_ip, subnet_sizes):
    try:
        # Vérifier si le masque est spécifié
        if '/' not in base_ip:
            raise ValueError("Vous devez spécifier un masque (ex: 10.0.0.0/8)")
            
        # Trier par taille décroissante (requis pour VLSM)
        subnet_sizes_sorted = sorted([s for s in subnet_sizes if s > 0], reverse=True)
        
        network = ipaddress.IPv4Network(base_ip, strict=False)
        subnets = []
        allocated_addresses = 0
        
        for i, hosts_needed in enumerate(subnet_sizes_sorted, 1):
            # Calculer la taille nécessaire (hôtes + réseau + broadcast)
            size_needed = hosts_needed + 2
            # Trouver le plus petit /N qui peut contenir cette taille
            prefixlen = 32 - (size_needed - 1).bit_length()
            
            # Calculer la taille du bloc
            block_size = 2**(32 - prefixlen)
            
            # Vérifier l'espace restant
            if allocated_addresses + block_size > network.num_addresses:
                required_prefix = 32 - (size_needed - 1).bit_length()
                raise ValueError(f"Espace insuffisant pour {hosts_needed} hôtes (besoin /{required_prefix})")
            
            # Créer le sous-réseau
            subnet_address = network.network_address + allocated_addresses
            subnet = ipaddress.IPv4Network(f"{subnet_address}/{prefixlen}", strict=False)
            
            subnet_info = {
                "id": i,
                "hosts_needed": hosts_needed,
                "network_address": str(subnet.network_address),
                "broadcast_address": str(subnet.broadcast_address),
                "first_host": str(subnet.network_address + 1),
                "last_host": str(subnet.broadcast_address - 1),
                "subnet_mask": str(subnet.netmask),
                "prefix": f"/{subnet.prefixlen}",
                "available_hosts": subnet.num_addresses - 2
            }
            
            subnets.append(subnet_info)
            allocated_addresses += block_size
        
        # Réorganiser dans l'ordre original
        original_order = []
        for size in subnet_sizes:
            for subnet in subnets:
                if subnet["hosts_needed"] == size and subnet not in original_order:
                    original_order.append(subnet)
                    break
        
        return original_order
    
    except Exception as e:
        print(f"\n\033[91mERREUR: {str(e)}\033[0m")
        print(f"Conseil: Vérifiez que votre réseau de base est assez large")
        print(f"Le réseau {base_ip} offre {network.num_addresses} adresses")
        print(f"Total nécessaire: {sum(2**(32 - (h+2).bit_length()) for h in subnet_sizes_sorted)}")
        return None

def export_to_excel(subnets, base_ip, filename="vlsm_results.xlsx"):
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "VLSM Results"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = openpyxl.styles.PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        center_alignment = Alignment(horizontal='center')
        
        # En-tête
        ws['A1'] = "Rapport VLSM - Adressage IP"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:H1')
        
        ws['A2'] = f"Réseau de base: {base_ip}"
        ws.merge_cells('A2:H2')
        
        ws['A3'] = f"Date: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws.merge_cells('A3:H3')
        
        # Entêtes
        headers = ["Sous-réseau", "Hôtes nécessaires", "Adresse réseau", "Masque", 
                  "Première IP", "Dernière IP", "Broadcast", "Hôtes disponibles"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
        
        # Données
        for row, subnet in enumerate(subnets, 6):
            ws.cell(row=row, column=1, value=f"Réseau {subnet['id']}").border = border
            ws.cell(row=row, column=2, value=subnet['hosts_needed']).border = border
            ws.cell(row=row, column=3, value=f"{subnet['network_address']}{subnet['prefix']}").border = border
            ws.cell(row=row, column=4, value=subnet['subnet_mask']).border = border
            ws.cell(row=row, column=5, value=subnet['first_host']).border = border
            ws.cell(row=row, column=6, value=subnet['last_host']).border = border
            ws.cell(row=row, column=7, value=subnet['broadcast_address']).border = border
            ws.cell(row=row, column=8, value=subnet['available_hosts']).border = border
        
        # Ajustement des colonnes
        for col in range(1, 9):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18
        
        wb.save(filename)
        print(f"\nRésultats enregistrés dans '{filename}'")
        
    except Exception as e:
        print(f"Erreur lors de l'export Excel: {e}")

def main():
    print("Calculateur VLSM - Adressage IP avec masque variable")
    print("----------------------------------------------------\n")
    
    base_ip = input("Entrez l'adresse réseau de base avec masque (ex: 10.0.0.0/16): ")
    num_subnets = int(input("Combien de sous-réseaux avez-vous besoin? "))
    
    subnet_sizes = []
    for i in range(num_subnets):
        while True:
            try:
                hosts = int(input(f"Nombre d'hôtes nécessaires pour le sous-réseau {i+1}: "))
                if hosts < 1:
                    print("Le nombre d'hôtes doit être ≥ 1")
                else:
                    subnet_sizes.append(hosts)
                    break
            except ValueError:
                print("Veuillez entrer un nombre valide")
    
    subnets = calculate_vlsm(base_ip, subnet_sizes)
    
    if subnets:
        print("\nRésultats de l'adressage VLSM:")
        print("="*60)
        for i, subnet in enumerate(subnets, 1):
            print(f"\nSous-réseau {i} (besoin: {subnet['hosts_needed']} hôtes):")
            print(f"Adresse réseau:    {subnet['network_address']}{subnet['prefix']}")
            print(f"Masque:            {subnet['subnet_mask']}")
            print(f"Plage d'adresses:  {subnet['first_host']} - {subnet['last_host']}")
            print(f"Broadcast:         {subnet['broadcast_address']}")
            print(f"Hôtes disponibles: {subnet['available_hosts']}")
        
        filename = input("\nNom du fichier Excel (Entrée pour vlsm_results.xlsx): ")
        if not filename:
            filename = "vlsm_results.xlsx"
        elif not filename.endswith('.xlsx'):
            filename += '.xlsx'
        
        export_to_excel(subnets, base_ip, filename)

if __name__ == "__main__":
    import openpyxl
    main()
